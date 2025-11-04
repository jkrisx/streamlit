import streamlit as st
from openpyxl import load_workbook
import io

uploaded_file = st.file_uploader("Upload Excel", type="xlsx")

if uploaded_file:
    # Convert file upload jadi BytesIO agar bisa dibaca openpyxl
    wb = load_workbook(io.BytesIO(uploaded_file.read()))
    ws = wb.active
    
    # Baca data dari cell
    st.write("Isi A1:", ws["A1"].value)
